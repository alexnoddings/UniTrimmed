from typing import Dict, List
from core.tool import Tool
from glob import glob
from core.aggregator import aggregate
from openpyxl import Workbook


class XlsxWriter(Tool):
    def __init__(self, args: Dict[str, str], working_dir: str = ""):
        super().__init__(args, working_dir)
        self.out_path = args["out"]

    def run(self, data: List[Dict]) -> List[Dict]:
        headers = ["URN"]
        for school in data:
            for key in school.keys():
                if key not in headers:
                    headers.append(key)

        workbook = Workbook()
        worksheet = workbook.active
        headers_map = {headers[i]: i + 1 for i in range(0, len(headers))}
        for i in range(0, len(headers)):
            worksheet.cell(row=1, column=i + 1, value=headers[i])

        for i in range(0, len(data)):
            school = data[i]
            for key, value in school.items():
                column = headers_map[key]
                worksheet.cell(row=i + 2, column=column, value=value)

        workbook.save(self.out_path)

        return data


entry = XlsxWriter
